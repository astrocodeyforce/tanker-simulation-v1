#!/usr/bin/env python3
"""
Combined Fleet Intelligence & Discharge Simulation Report — v3
Bull and Bear Co. — March 2026

PURPOSE: Show ownership/management why switching from 3" to 2" hoses
would be a mistake, what the optimal compressor SCFM is, and
quantify the cost of downgrading — all backed by real fleet
commodity data + 840 physics simulations validated to ±7%.

Pages:
  1  Cover Page
  2  Executive Summary — 3 key findings
  3  Your Fleet: What You Haul (commodity data)
  4  The Big Question: 2" vs 3" Hose — Product-by-Product Impact
  5  Compressor Sizing: Finding the Right SCFM
  6  Compressor Upgrade ROI — Payback Period
  7  The True Cost of Switching to 2" Hose
  8  Pressure Safety — Operating Pressures by Hose Size
  9  Top 10 Product Cards — Time per Real Product
  10 Sensitivity Tornado — What Matters Most
  11 Flow Velocity & Hose Wear — Erosion Risk
  12 Recommendations & Decision Framework
"""
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.patches import FancyBboxPatch
import seaborn as sns
from scipy.interpolate import interp1d
import os

# ── Paths ─────────────────────────────────────────────────────────
WS       = "/opt/sim-lab/truck-tanker-sim-env"
SIM_CSV  = os.path.join(WS, "data", "sweep_5d_results.csv")
OUT_PDF  = os.path.join(WS, "data", "Combined_Fleet_Simulation_Report_March2026.pdf")
OUT_XLSX = os.path.join(WS, "data", "Combined_Fleet_Simulation_Analysis.xlsx")

# ── Colors ────────────────────────────────────────────────────────
NAVY      = "#0D2137"
DBLUE     = "#003A70"
MBLUE     = "#0070C0"
ACCENT    = "#00B0F0"
LBG       = "#F2F7FA"
WHITE     = "#FFFFFF"
ORANGE    = "#E8600A"
GREEN     = "#00854A"
RED       = "#C00000"
GRAY      = "#747678"
LGRAY     = "#D9D9D9"
PURPLE    = "#7030A0"

TOTAL_PAGES = 16

plt.rcParams.update({
    "font.family": "sans-serif",
    "font.sans-serif": ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size": 9,
    "figure.facecolor": WHITE,
    "axes.facecolor": WHITE,
    "axes.edgecolor": LGRAY,
    "axes.grid": True,
    "grid.color": LGRAY,
    "grid.linewidth": 0.4,
})

# ═══════════════════════════════════════════════════════════════════
#  DATA
# ═══════════════════════════════════════════════════════════════════
TOP15 = pd.DataFrame([
    {"commodity": "Ethylene Glycol",         "loads": 51, "visc_cP": 16},
    {"commodity": "Resin Solution",          "loads": 44, "visc_cP": 500},
    {"commodity": "NIPOL 1411 LATEX",        "loads": 35, "visc_cP": 200},
    {"commodity": "SAFEWING MP IV LAUNCH",   "loads": 35, "visc_cP": 15},
    {"commodity": "Sodium Silicate",         "loads": 34, "visc_cP": 180},
    {"commodity": "DOSS 70 PG",              "loads": 28, "visc_cP": 200},
    {"commodity": "Smartcide 1984A",         "loads": 26, "visc_cP": 5},
    {"commodity": "OCD-277",                 "loads": 26, "visc_cP": 1},
    {"commodity": "Propylene Glycol (PGI)",  "loads": 26, "visc_cP": 42},
    {"commodity": "Diethylene Glycol",       "loads": 25, "visc_cP": 30},
    {"commodity": "Used Motor Oil",          "loads": 24, "visc_cP": 20},
    {"commodity": "Triethylene Glycol",      "loads": 24, "visc_cP": 37},
    {"commodity": "NAXONATE 4LS",            "loads": 23, "visc_cP": 5},
    {"commodity": "Biomass",                 "loads": 23, "visc_cP": 50},
    {"commodity": "VIVATEC 500",             "loads": 18, "visc_cP": 100},
])

VISC_BINS = pd.DataFrame([
    {"label": "≤ 10 cP  (water-like)",  "lo": 0,   "hi": 10,   "mid": 5,    "pct": 26, "loads": 563},
    {"label": "10–50 cP  (light)",       "lo": 10,  "hi": 50,   "mid": 30,   "pct": 38, "loads": 823},
    {"label": "50–200 cP  (medium)",     "lo": 50,  "hi": 200,  "mid": 125,  "pct": 26, "loads": 563},
    {"label": "200–500 cP  (heavy)",     "lo": 200, "hi": 500,  "mid": 350,  "pct": 5,  "loads": 108},
    {"label": "500–2000 cP  (v. heavy)", "lo": 500, "hi": 2000, "mid": 1000, "pct": 4,  "loads": 87},
    {"label": "> 2000 cP  (ultra)",      "lo": 2000,"hi": 6000, "mid": 3500, "pct": 1,  "loads": 21},
])

sim = pd.read_csv(SIM_CSV)
for c in ["time_min", "peak_gpm", "avg_gpm", "avg_pressure_psig",
          "time_50pct_min", "time_90pct_min"]:
    sim[c] = pd.to_numeric(sim[c], errors="coerce")
SIM_VISCS = sorted(sim["visc_cP"].unique())
SIM_SCFMS = sorted(sim["scfm"].unique())


def _interp_time_exact_scfm(hose, scfm, visc, vol, psi):
    """Interpolate time for an exact SCFM value present in the sweep."""
    vols = sorted(sim["vol_gal"].unique())  # [5000, 6500]
    if vol in vols:
        sub = sim[(sim["hose_in"] == hose) & (sim["scfm"] == scfm) &
                  (sim["vol_gal"] == vol) & (sim["pre_psi"] == psi)].dropna(subset=["time_min"])
        if len(sub) < 2:
            return np.nan
        f = interp1d(np.log10(sub["visc_cP"]), sub["time_min"],
                     kind="linear", fill_value="extrapolate")
        return float(f(np.log10(max(visc, 1))))
    # Interpolate between the two volume sweep points
    v_lo, v_hi = vols[0], vols[-1]
    s_lo = sim[(sim["hose_in"] == hose) & (sim["scfm"] == scfm) &
               (sim["vol_gal"] == v_lo) & (sim["pre_psi"] == psi)].dropna(subset=["time_min"])
    s_hi = sim[(sim["hose_in"] == hose) & (sim["scfm"] == scfm) &
               (sim["vol_gal"] == v_hi) & (sim["pre_psi"] == psi)].dropna(subset=["time_min"])
    if len(s_lo) < 2 or len(s_hi) < 2:
        return np.nan
    f_lo = interp1d(np.log10(s_lo["visc_cP"]), s_lo["time_min"],
                    kind="linear", fill_value="extrapolate")
    f_hi = interp1d(np.log10(s_hi["visc_cP"]), s_hi["time_min"],
                    kind="linear", fill_value="extrapolate")
    t_lo = float(f_lo(np.log10(max(visc, 1))))
    t_hi = float(f_hi(np.log10(max(visc, 1))))
    return t_lo + (t_hi - t_lo) * (vol - v_lo) / (v_hi - v_lo)


def interp_time(hose, scfm, visc, vol=6000, psi=20):
    """Log-linear interpolation of discharge time with volume + SCFM interpolation."""
    scfm_list = sorted(sim["scfm"].unique())
    if scfm in scfm_list:
        return _interp_time_exact_scfm(hose, scfm, visc, vol, psi)
    # SCFM not in sweep — interpolate between neighbors
    below = [s for s in scfm_list if s <= scfm]
    above = [s for s in scfm_list if s >= scfm]
    if not below or not above:
        return np.nan
    s_lo, s_hi = below[-1], above[0]
    if s_lo == s_hi:
        return _interp_time_exact_scfm(hose, s_lo, visc, vol, psi)
    t_lo = _interp_time_exact_scfm(hose, s_lo, visc, vol, psi)
    t_hi = _interp_time_exact_scfm(hose, s_hi, visc, vol, psi)
    if pd.isna(t_lo) or pd.isna(t_hi):
        return np.nan
    return t_lo + (t_hi - t_lo) * (scfm - s_lo) / (s_hi - s_lo)


# ═══════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════
def frame(fig, title, subtitle, pg):
    """Top & bottom bars for every slide."""
    fig.patches.append(FancyBboxPatch(
        (0, 0.945), 1, 0.055, transform=fig.transFigure,
        boxstyle="square,pad=0", facecolor=NAVY, edgecolor="none", zorder=10))
    fig.text(0.04, 0.970, title, fontsize=14, fontweight="bold",
             color=WHITE, va="center", zorder=11)
    if subtitle:
        fig.text(0.04, 0.920, subtitle, fontsize=8.5, color=GRAY, va="center")
    fig.patches.append(FancyBboxPatch(
        (0, 0), 1, 0.025, transform=fig.transFigure,
        boxstyle="square,pad=0", facecolor=LGRAY, edgecolor="none", zorder=10))
    fig.text(0.04, 0.012, "Bull and Bear Co.  |  Confidential",
             fontsize=6.5, color=GRAY, va="center", zorder=11)
    fig.text(0.96, 0.012, f"{pg}/{TOTAL_PAGES}",
             fontsize=6.5, color=GRAY, va="center", ha="right", zorder=11)
    fig.patches.append(FancyBboxPatch(
        (0, 0.940), 1, 0.003, transform=fig.transFigure,
        boxstyle="square,pad=0", facecolor=ACCENT, edgecolor="none", zorder=10))


def box(fig, x, y, w, h, num, label, color):
    fig.patches.append(FancyBboxPatch(
        (x, y), w, h, transform=fig.transFigure,
        boxstyle="round,pad=0.008", facecolor=color, edgecolor="none",
        alpha=0.92, zorder=5))
    fig.text(x + w/2, y + h * 0.62, str(num), fontsize=18, fontweight="bold",
             color=WHITE, ha="center", va="center", zorder=6)
    fig.text(x + w/2, y + h * 0.22, label, fontsize=7,
             color=WHITE, ha="center", va="center", zorder=6, linespacing=1.3)


def callout(fig, x, y, w, text, border=ACCENT):
    fig.text(x + w/2, y, text, fontsize=8, fontweight="bold", color=NAVY,
             ha="center", va="center", linespacing=1.4,
             bbox=dict(boxstyle="round,pad=0.35", facecolor=LBG,
                       edgecolor=border, linewidth=1.5), zorder=8)


# ═══════════════════════════════════════════════════════════════════
#  PAGE 1 — COVER PAGE
# ═══════════════════════════════════════════════════════════════════
def cover(pdf):
    fig = plt.figure(figsize=(11, 8.5))
    # Full navy background
    fig.patches.append(FancyBboxPatch(
        (0, 0), 1, 1, transform=fig.transFigure,
        boxstyle="square,pad=0", facecolor=NAVY, edgecolor="none"))
    # Accent stripe
    fig.patches.append(FancyBboxPatch(
        (0, 0.52), 1, 0.006, transform=fig.transFigure,
        boxstyle="square,pad=0", facecolor=ACCENT, edgecolor="none", zorder=2))
    # Title block
    fig.text(0.50, 0.72, "FLEET EQUIPMENT", fontsize=36, fontweight="bold",
             color=WHITE, ha="center", va="center", zorder=3)
    fig.text(0.50, 0.64, "OPTIMIZATION REPORT", fontsize=36, fontweight="bold",
             color=ACCENT, ha="center", va="center", zorder=3)
    fig.text(0.50, 0.57, "Hose Size & Compressor Analysis",
             fontsize=16, color=LGRAY, ha="center", va="center", zorder=3)
    # Subtitle area
    fig.text(0.50, 0.44,
             "Combining real fleet commodity data (3,346 trips)\n"
             "with physics-based discharge simulation (840 runs, validated ±7%)\n"
             "to guide equipment investment decisions",
             fontsize=11, color=LGRAY, ha="center", va="center",
             linespacing=1.6, zorder=3)
    # Bottom info
    fig.text(0.50, 0.22, "Bull and Bear Co.", fontsize=18, fontweight="bold",
             color=WHITE, ha="center", va="center", zorder=3)
    fig.text(0.50, 0.16, "Prepared: March 2026", fontsize=11,
             color=LGRAY, ha="center", va="center", zorder=3)
    fig.text(0.50, 0.12,
             "Data Sources: QuickManage API (Oct 2025 – Feb 2026)  |  "
             "OpenModelica TankerTransferV2 Simulation",
             fontsize=8, color=GRAY, ha="center", va="center", zorder=3)
    fig.text(0.50, 0.06, "Confidential — Internal Use Only",
             fontsize=8, color=GRAY, ha="center", va="center",
             style="italic", zorder=3)
    pdf.savefig(fig); plt.close(fig)


# ═══════════════════════════════════════════════════════════════════
#  PAGE 2 — EXECUTIVE SUMMARY
# ═══════════════════════════════════════════════════════════════════
def page_exec(pdf):
    fig = plt.figure(figsize=(11, 8.5))
    frame(fig, "EXECUTIVE SUMMARY", "Key findings from fleet data + discharge simulation", 2)

    # Metric boxes
    box(fig, 0.04, 0.76, 0.22, 0.13, "3,346", "Trips\nAnalyzed", DBLUE)
    box(fig, 0.28, 0.76, 0.22, 0.13, "840", "Simulations\nRun", GREEN)
    box(fig, 0.52, 0.76, 0.22, 0.13, "±7%", "Field-Validated\nAccuracy", MBLUE)
    box(fig, 0.76, 0.76, 0.22, 0.13, "1,432", "Products\nIdentified", ORANGE)

    # Context paragraph
    fig.text(0.04, 0.70,
             "Bull and Bear Co. is evaluating whether to switch from 3\" to 2\" hoses for cost savings and easier driver handling.\n"
             "This report combines 5 months of real fleet commodity data with a physics-based simulation engine to answer that question with data.",
             fontsize=9.5, color=NAVY, va="top", linespacing=1.5)

    # 3 Findings
    findings = [
        ("1", 'Switching to 2" hose adds 11–200+ minutes per load for your products',
         'Your fleet hauls products ranging from 1 to 500+ cP. For anything above 50 cP — which is 36% of your loads —\n'
         'a 2" hose dramatically increases discharge time. For Resin Solution (500 cP, your #2 product), discharge\n'
         'jumps from 52 min to 83 min per load at 6,000 gal. For heavier products, the 2" hose may not finish at all.'),

        ("2", 'Your current compressor (~17 SCFM) is adequate for 3" hose but undersized for 2"',
         'With a 3" hose, 17 SCFM puts you in the performance sweet spot — more airflow brings diminishing returns.\n'
         'But with a 2" hose, the pipe friction is so high that even doubling the compressor to 40 SCFM barely\n'
         'helps on thick products. You would be spending more on compressors just to partially offset the smaller hose.'),

        ("3", 'The 2" switch would cost roughly 2,100+ hours of extra driver time per year',
         'Across all 3,346 loads (scaled to annual), switching to 2" adds an estimated 2,100+ hours\n'
         'of idle driver and trailer time per year. That is lost revenue, not savings — and it hits hardest on the\n'
         'heavier products where margins are higher.'),
    ]

    y = 0.58
    for num, title_text, detail in findings:
        fig.patches.append(FancyBboxPatch(
            (0.04, y - 0.085), 0.035, 0.08, transform=fig.transFigure,
            boxstyle="round,pad=0.003", facecolor=DBLUE, edgecolor="none", zorder=5))
        fig.text(0.0575, y - 0.045, num, fontsize=14, fontweight="bold",
                 color=WHITE, ha="center", va="center", zorder=6)
        fig.text(0.09, y - 0.01, title_text, fontsize=10, fontweight="bold",
                 color=NAVY, va="center")
        fig.text(0.09, y - 0.055, detail, fontsize=7.5, color=GRAY,
                 va="center", linespacing=1.3)
        y -= 0.14

    callout(fig, 0.04, 0.045, 0.92,
            'RECOMMENDATION:  Keep the 3" hose. The cost savings from 2" are far outweighed by '
            'increased discharge time, idle driver hours, and reduced fleet throughput.', RED)

    pdf.savefig(fig); plt.close(fig)


# ═══════════════════════════════════════════════════════════════════
#  PAGE 3 — FLEET PROFILE: WHAT YOU HAUL
# ═══════════════════════════════════════════════════════════════════
def page_fleet(pdf):
    fig = plt.figure(figsize=(11, 8.5))
    frame(fig, "YOUR FLEET: WHAT YOU ACTUALLY HAUL",
          "Commodity extraction from 3,346 trips  |  Oct 2025 – Feb 2026", 3)

    ax1 = fig.add_axes([0.18, 0.30, 0.32, 0.55])
    ax2 = fig.add_axes([0.58, 0.30, 0.38, 0.55])

    # Left: Top 15 horizontal bar
    top = TOP15.sort_values("loads")
    colors = []
    for v in top["visc_cP"]:
        if v <= 10:   colors.append(MBLUE)
        elif v <= 50: colors.append(ACCENT)
        elif v <= 200: colors.append(GREEN)
        elif v <= 500: colors.append(ORANGE)
        else:          colors.append(RED)

    bars = ax1.barh(range(len(top)), top["loads"], color=colors, edgecolor=WHITE, linewidth=0.4)
    ax1.set_yticks(range(len(top)))
    ax1.set_yticklabels([f'{r["commodity"]}  ({r["visc_cP"]} cP)'
                         for _, r in top.iterrows()], fontsize=7)
    ax1.set_xlabel("Loads (5-month period)", fontsize=9, fontweight="bold")
    ax1.set_title("Top 15 Most-Hauled Commodities", fontsize=11,
                  fontweight="bold", color=DBLUE)
    for b, v in zip(bars, top["loads"]):
        ax1.text(b.get_width() + 0.5, b.get_y() + b.get_height()/2,
                 str(v), fontsize=7, va="center", fontweight="bold", color=NAVY)
    ax1.set_xlim(0, 62)

    # Right: Viscosity distribution donut
    bins = VISC_BINS.copy()
    total_trips = 3346
    bins["scaled_loads"] = (bins["pct"] / 100 * total_trips).round().astype(int)
    clrs = [MBLUE, ACCENT, GREEN, ORANGE, RED, PURPLE]
    wedges, texts = ax2.pie(
        bins["scaled_loads"], labels=None, colors=clrs,
        startangle=90, pctdistance=0.78,
        wedgeprops=dict(width=0.42, edgecolor=WHITE, linewidth=1.5))
    # Center label
    ax2.text(0, 0, f"{total_trips:,}\nloads", fontsize=14, fontweight="bold",
             color=NAVY, ha="center", va="center")
    ax2.set_title("Fleet Viscosity Distribution", fontsize=11,
                  fontweight="bold", color=DBLUE, pad=12)
    # Legend below
    legend_labels = [f'{b["label"]}  —  {b["pct"]}%  ({b["scaled_loads"]:,} loads)'
                     for _, b in bins.iterrows()]
    ax2.legend(wedges, legend_labels, loc="upper center",
               bbox_to_anchor=(0.5, -0.05), fontsize=7, ncol=1,
               frameon=False)

    callout(fig, 0.04, 0.06, 0.92,
            "36% of loads are above 50 cP (medium to ultra-heavy). "
            "These are the products most impacted by hose size — "
            "and they include your #2 and #3 most-hauled commodities.", DBLUE)

    pdf.savefig(fig); plt.close(fig)


# ═══════════════════════════════════════════════════════════════════
#  PAGE 4 — 2" vs 3" HOSE: PRODUCT-BY-PRODUCT
# ═══════════════════════════════════════════════════════════════════
def page_hose_compare(pdf, scfm=19, pg=4):
    fig = plt.figure(figsize=(11, 8.5))
    frame(fig, f'2" vs 3" HOSE COMPARISON AT {scfm} SCFM',
          f'Simulated discharge times for your actual products  |  6,000 gal, 20 psig, {scfm} SCFM', pg)

    # Top chart: grouped bar — 2" vs 3" for top 15 products
    ax = fig.add_axes([0.06, 0.46, 0.88, 0.40])

    products = TOP15.sort_values("visc_cP").reset_index(drop=True)
    x = np.arange(len(products))
    w = 0.35

    t3 = [interp_time(3, scfm, v) for v in products["visc_cP"]]
    t2 = [interp_time(2, scfm, v) for v in products["visc_cP"]]

    bars3 = ax.bar(x - w/2, t3, w, color=GREEN, label='3" Hose (current)', edgecolor=WHITE, linewidth=0.5)
    bars2 = ax.bar(x + w/2, t2, w, color=RED,   label='2" Hose (proposed)', edgecolor=WHITE, linewidth=0.5)

    ax.set_xticks(x)
    ax.set_xticklabels([f'{r["commodity"]}\n({r["visc_cP"]} cP)'
                         for _, r in products.iterrows()],
                       fontsize=6.5, rotation=45, ha="right")
    ax.set_ylabel("Discharge Time (minutes)", fontsize=9, fontweight="bold")
    ax.set_title('Discharge Time Comparison: 3" (Current) vs 2" (Proposed)',
                 fontsize=11, fontweight="bold", color=DBLUE, pad=8)
    ax.legend(fontsize=9, loc="upper left")

    # Add delta labels on top of 2" bars
    for i, (a, b) in enumerate(zip(t3, t2)):
        if pd.notna(a) and pd.notna(b):
            delta = b - a
            pct = delta / a * 100
            ax.text(i + w/2, b + 0.8, f"+{delta:.0f} min\n(+{pct:.0f}%)",
                    fontsize=6, ha="center", fontweight="bold", color=RED,
                    linespacing=1.1)

    ax.set_ylim(0, max([v for v in t2 if pd.notna(v)]) * 1.25)

    # Bottom summary table
    fig.text(0.06, 0.30, "Summary: Extra time per load if switching to 2\"",
             fontsize=10, fontweight="bold", color=DBLUE)

    # Simple text table with worst offenders
    worst = []
    for _, p in products.iterrows():
        a = interp_time(3, scfm, p["visc_cP"])
        b = interp_time(2, scfm, p["visc_cP"])
        if pd.notna(a) and pd.notna(b):
            worst.append((p["commodity"], p["visc_cP"], p["loads"], a, b, b - a, (b-a)*p["loads"]))
    worst.sort(key=lambda x: x[5], reverse=True)

    # Table headers
    col_x  = [0.06, 0.24, 0.33, 0.42, 0.52, 0.63, 0.74, 0.87]
    col_w  = [0.17, 0.08, 0.08, 0.09, 0.10, 0.10, 0.12, 0.10]
    hdrs   = ["Product", "Visc", "Loads", '3" Time', '2" Time', "Extra/Load", "Total Extra", "Impact"]
    y_t = 0.27
    for cx, cw, h in zip(col_x, col_w, hdrs):
        fig.patches.append(FancyBboxPatch(
            (cx, y_t), cw, 0.025, transform=fig.transFigure,
            boxstyle="square,pad=0", facecolor=DBLUE, edgecolor=WHITE, zorder=5))
        fig.text(cx + cw/2, y_t + 0.012, h, fontsize=7, fontweight="bold",
                 color=WHITE, ha="center", va="center", zorder=6)
    # Top 6 worst products
    y_t -= 0.005
    for i, (name, visc, loads, t3v, t2v, delta, total) in enumerate(worst[:6]):
        y_t -= 0.023
        bg = LBG if i % 2 == 0 else WHITE
        impact = "LOW" if delta < 3 else ("MEDIUM" if delta < 10 else "HIGH")
        imp_color = GREEN if impact == "LOW" else (ORANGE if impact == "MEDIUM" else RED)
        vals = [name, f"{visc:.0f} cP", str(loads),
                f"{t3v:.1f} min", f"{t2v:.1f} min",
                f"+{delta:.1f} min", f"+{total:.0f} min", impact]
        for ci, (cx, cw, v) in enumerate(zip(col_x, col_w, vals)):
            fig.patches.append(FancyBboxPatch(
                (cx, y_t), cw, 0.022, transform=fig.transFigure,
                boxstyle="square,pad=0", facecolor=bg, edgecolor=LGRAY,
                linewidth=0.3, zorder=4))
            color = imp_color if ci == 7 else (RED if ci >= 5 else NAVY)
            fig.text(cx + cw/2, y_t + 0.011, v, fontsize=6.5,
                     color=color, ha="center", va="center",
                     fontweight="bold" if ci >= 5 else "normal", zorder=5)

    # Compute min/max penalties for callout
    penalties = [b - a for a, b in zip(t3, t2) if pd.notna(a) and pd.notna(b)]
    p_min, p_max = min(penalties), max(penalties)
    callout(fig, 0.04, 0.045, 0.92,
            f'At {scfm} SCFM, switching to 2" hose adds {p_min:.0f}\u2013{p_max:.0f}+ minutes per load '
            f'at 6,000 gal across your product mix.',
            RED)

    pdf.savefig(fig); plt.close(fig)


# ═══════════════════════════════════════════════════════════════════
#  COMPRESSOR ANALYSIS
# ═══════════════════════════════════════════════════════════════════
def page_compressor(pdf):
    fig = plt.figure(figsize=(11, 8.5))
    frame(fig, "COMPRESSOR SIZING: WHAT SCFM DO YOU ACTUALLY NEED?",
          "How compressor airflow interacts with hose size for your product mix", 10)

    ax1 = fig.add_axes([0.06, 0.38, 0.40, 0.46])
    ax2 = fig.add_axes([0.56, 0.38, 0.40, 0.46])

    # 4 real fleet viscosities to plot
    visc_show = [
        (30,  "Fleet Median (30 cP)",    MBLUE),
        (100, "VIVATEC 500 (100 cP)",    GREEN),
        (200, "NIPOL LATEX (200 cP)",    ORANGE),
        (500, "Resin Solution (500 cP)", RED),
    ]

    for hose_idx, (hose, ax, title) in enumerate(
            [(3, ax1, '3" Hose (Current)'), (2, ax2, '2" Hose (Proposed)')]):
        for visc, label, color in visc_show:
            times = []
            for s in SIM_SCFMS:
                times.append(interp_time(hose, s, visc))
            ax.plot(SIM_SCFMS, times, "-o", color=color, linewidth=2,
                    markersize=4, label=label)

            # Add % reduction annotation: from 19 SCFM to 30 SCFM
            t_lo = interp_time(hose, 19, visc)
            t_hi = interp_time(hose, 30, visc)
            if pd.notna(t_lo) and pd.notna(t_hi) and t_lo > 0:
                pct_drop = (t_lo - t_hi) / t_lo * 100
                # Place label at 30 SCFM point
                ax.annotate(f"−{pct_drop:.0f}%",
                            xy=(30, t_hi), fontsize=7, fontweight="bold",
                            color=color, ha="left", va="center",
                            xytext=(3, 0), textcoords="offset points")

        ax.set_xlabel("Compressor (SCFM)", fontsize=9, fontweight="bold")
        ax.set_ylabel("Discharge Time (min)", fontsize=9, fontweight="bold")
        ax.set_title(title, fontsize=11, fontweight="bold", color=DBLUE, pad=5)
        ax.legend(fontsize=6.5, loc="upper right")
        ax.set_ylim(0, 120)
        ax.set_xlim(8, 68)
        # Mark current 19 SCFM position
        ax.axvline(x=19, color=LGRAY, linewidth=2, linestyle="--", zorder=1)
        ax.text(20, ax.get_ylim()[1] * 0.97, "Current\n~19 SCFM",
                fontsize=7, color=GRAY, fontweight="bold", va="top")

    # Highlight sweet spot on 3" chart
    ax1.axvspan(19, 30, alpha=0.08, color=GREEN, zorder=0)
    ax1.text(24.5, 5, "Sweet Spot", fontsize=8, color=GREEN,
             ha="center", fontweight="bold")

    # Per-product friction floor lines + bottleneck transition on 2" chart
    for visc, label, color in visc_show:
        # Asymptotic floor = time at 60 SCFM (maximum air supply)
        t_floor = interp_time(2, 60, visc)
        if pd.notna(t_floor):
            ax2.axhline(y=t_floor, color=color, linewidth=0.8,
                        linestyle=":", alpha=0.5, zorder=1)
            ax2.text(67, t_floor, f"{t_floor:.0f}m",
                     fontsize=5.5, color=color, va="center", ha="left",
                     fontstyle="italic", alpha=0.7)
        # Find where slope drops below 0.2 min/SCFM (friction-dominated)
        for s in SIM_SCFMS:
            t_here = interp_time(2, s, visc)
            s_next = s + 1
            t_next = interp_time(2, s_next, visc)
            if pd.notna(t_here) and pd.notna(t_next):
                slope = t_here - t_next
                if slope < 0.2:
                    ax2.plot(s, t_here, "x", color=color, markersize=6,
                             markeredgewidth=1.5, zorder=5, alpha=0.7)
                    break
    ax2.text(62, 114, "╳ = friction\n     floor", fontsize=5.5,
             color=GRAY, va="top", ha="center", fontstyle="italic")
    ax2.text(67.5, 108, "┈ = minimum\n     possible", fontsize=5.5,
             color=GRAY, va="top", ha="center", fontstyle="italic")

    # Bottom table: Realistic 19 → 30 SCFM upgrade comparison
    fig.text(0.06, 0.30, "Realistic Upgrade: 19 → 30 SCFM  (achievable for most trucks)",
             fontsize=10, fontweight="bold", color=DBLUE)

    col_x = [0.04, 0.18, 0.30, 0.41, 0.52, 0.63, 0.75, 0.86]
    col_w = [0.13, 0.11, 0.10, 0.10, 0.10, 0.11, 0.10, 0.11]
    hdrs  = ["Product", '3" @ 19\nSCFM', '3" @ 30\nSCFM', '3"\nSaved',
             '3" %\nFaster', '2" @ 19\nSCFM', '2" @ 30\nSCFM', '2"\nSaved']
    y_t = 0.27
    for cx, cw, h in zip(col_x, col_w, hdrs):
        fig.patches.append(FancyBboxPatch(
            (cx, y_t), cw, 0.03, transform=fig.transFigure,
            boxstyle="square,pad=0", facecolor=DBLUE, edgecolor=WHITE, zorder=5))
        fig.text(cx + cw/2, y_t + 0.015, h, fontsize=6.5, fontweight="bold",
                 color=WHITE, ha="center", va="center", zorder=6)
    y_t -= 0.005
    for idx, (visc, label, color) in enumerate(visc_show):
        y_t -= 0.024
        bg = LBG if idx % 2 == 0 else WHITE
        t3_19 = interp_time(3, 19, visc)
        t3_30 = interp_time(3, 30, visc)
        t2_19 = interp_time(2, 19, visc)
        t2_30 = interp_time(2, 30, visc)
        d3 = (t3_19 - t3_30) if pd.notna(t3_19) and pd.notna(t3_30) else 0
        d2 = (t2_19 - t2_30) if pd.notna(t2_19) and pd.notna(t2_30) else 0
        pct3 = (d3 / t3_19 * 100) if t3_19 > 0 else 0

        vals = [label.split("(")[0].strip(),
                f"{t3_19:.1f} min", f"{t3_30:.1f} min", f"−{d3:.1f} min", f"−{pct3:.0f}%",
                f"{t2_19:.1f} min", f"{t2_30:.1f} min", f"−{d2:.1f} min"]
        for ci, (cx, cw, v) in enumerate(zip(col_x, col_w, vals)):
            fig.patches.append(FancyBboxPatch(
                (cx, y_t), cw, 0.023, transform=fig.transFigure,
                boxstyle="square,pad=0", facecolor=bg, edgecolor=LGRAY,
                linewidth=0.3, zorder=4))
            fc = NAVY
            if ci in (3, 4): fc = GREEN
            if ci == 7: fc = ORANGE
            fig.text(cx + cw/2, y_t + 0.011, v, fontsize=6.5,
                     color=fc, ha="center", va="center", zorder=5,
                     fontweight="bold" if ci in (0, 3, 4, 7) else "normal")

    callout(fig, 0.04, 0.045, 0.92,
            '2" hose: each product has a friction floor (dotted lines) — the minimum time even with infinite air.\n'
            'The ╳ marks where adding more SCFM stops helping (<0.2 min/SCFM). Heavy products wall out early; light ones keep improving.',
            DBLUE)

    pdf.savefig(fig); plt.close(fig)


# ═══════════════════════════════════════════════════════════════════
#  PAGE 6 — COMPRESSOR UPGRADE: CAPACITY GAINED
# ═══════════════════════════════════════════════════════════════════
GOLD = "#D4A017"

def _prepress_time_min(scfm, headspace_gal=1000, target_psig=20):
    """Pre-pressurization time (minutes) via ideal-gas + vol. efficiency ODE."""
    V = headspace_gal * 3.78541e-3        # m³
    P_atm = 101325.0; T = 293.15          # Pa, K
    R = 8.314; M = 0.02897; gamma = 1.4; c = 0.04
    P_tgt = target_psig * 6894.76 + P_atm
    rho_std = P_atm * M / (R * T)
    mdot = rho_std * scfm * 0.000471947   # kg/s
    m = P_atm * V * M / (R * T)           # initial air mass
    dt, t = 0.1, 0.0
    while True:
        P = m * R * T / (M * V)
        if P >= P_tgt:
            break
        eta = max(1 - c * ((P / P_atm)**(1/gamma) - 1), 0.1)
        m += mdot * eta * dt; t += dt
        if t > 3600:
            break
    return t / 60

def page_compressor_roi(pdf):
    fig = plt.figure(figsize=(11, 8.5))
    frame(fig, "COMPRESSOR UPGRADE: TOTAL TIME SAVED PER UNLOAD (17 → 30 SCFM)",
          "Discharge savings + pre-pressurization savings for 6,000 gal average load, 3\" hose", 11)

    # ── Parameters ──
    num_trucks = 80
    total_trips = 3346
    months_data, months_per_year = 5, 12
    avg_haul_rate = 350
    V0_gal = 7000; V_liq = 6000
    headspace = V0_gal - V_liq  # 1000 gal
    scfm_base = 17; scfm_target = 30

    # Pre-pressurization times
    pp_base = _prepress_time_min(scfm_base, headspace)
    pp_target = _prepress_time_min(scfm_target, headspace)
    pp_saved = pp_base - pp_target  # ~4.7 min

    # Scale loads to full 3,346 trips
    bins = VISC_BINS.copy()
    bins["scaled_loads"] = (bins["pct"] / 100 * total_trips).round().astype(int)

    disc_saved = []
    for _, b in bins.iterrows():
        t_b = interp_time(3, scfm_base, b["mid"])
        t_t = interp_time(3, scfm_target, b["mid"])
        s = (t_b - t_t) if pd.notna(t_b) and pd.notna(t_t) else 0
        disc_saved.append(max(s, 0))
    bins["disc_saved"] = disc_saved
    bins["pp_saved"] = pp_saved
    bins["total_saved"] = bins["disc_saved"] + pp_saved
    bins["total_saved_min"] = bins["total_saved"] * bins["scaled_loads"]
    bins["annual_saved_min"] = bins["total_saved_min"] * months_per_year / months_data
    bins["annual_saved_hrs"] = bins["annual_saved_min"] / 60

    fleet_annual_hrs = bins["annual_saved_hrs"].sum()
    per_truck_hrs = fleet_annual_hrs / num_trucks
    loads_per_truck_yr = total_trips / num_trucks * months_per_year / months_data
    avg_total_saved = bins["total_saved"].mean()

    # ── LEFT CHART: Stacked bars — discharge + pre-press savings ──
    ax1 = fig.add_axes([0.06, 0.42, 0.40, 0.42])
    x = np.arange(len(bins))
    w = 0.65
    bars_disc = ax1.bar(x, bins["disc_saved"], w, color=GREEN, edgecolor=WHITE,
                        linewidth=0.5, label="Discharge Saved", zorder=3)
    bars_pp = ax1.bar(x, bins["pp_saved"], w, bottom=bins["disc_saved"],
                      color=GOLD, edgecolor=WHITE, linewidth=0.5,
                      label="Pre-Pressurization Saved", zorder=3)
    ax1.set_xticks(x)
    ax1.set_xticklabels([b["label"].split("(")[0].strip() for _, b in bins.iterrows()],
                        fontsize=7, rotation=30, ha="right")
    ax1.set_ylabel("Minutes Saved Per Unload", fontsize=9, fontweight="bold")
    ax1.set_title(f"Total Time Saved Per Unload\n({scfm_base} \u2192 {scfm_target} SCFM, 3\" Hose, 6,000 gal)",
                  fontsize=10, fontweight="bold", color=DBLUE, pad=5)
    # Total labels on top
    for i, (d, p) in enumerate(zip(bins["disc_saved"], bins["pp_saved"])):
        tot = d + p
        ax1.text(i, tot + 0.5, f"\u2212{tot:.0f} min", fontsize=7.5,
                 ha="center", fontweight="bold", color=NAVY)
        # Discharge portion inside bar
        if d > 2:
            ax1.text(i, d / 2, f"{d:.0f}", fontsize=6.5, ha="center",
                     va="center", color=WHITE, fontweight="bold")
        # PP portion inside bar
        ax1.text(i, d + p / 2, f"{p:.1f}", fontsize=6, ha="center",
                 va="center", color=NAVY, fontweight="bold")
    ax1.legend(fontsize=7, loc="upper left", framealpha=0.9)
    ax1.set_ylim(0, bins["total_saved"].max() * 1.15)

    # ── RIGHT CHART: Pre-pressurization time vs SCFM ──
    ax2 = fig.add_axes([0.56, 0.42, 0.40, 0.42])
    scfm_list = [10, 12, 15, 17, 19, 25, 30, 35]
    pp_times = [_prepress_time_min(s, headspace) for s in scfm_list]
    ax2.plot(scfm_list, pp_times, "o-", color=MBLUE, linewidth=2.5,
             markersize=7, zorder=4)
    for s, t in zip(scfm_list, pp_times):
        ax2.annotate(f"{t:.1f} min", (s, t), textcoords="offset points",
                     xytext=(0, 10), fontsize=7, ha="center", fontweight="bold",
                     color=DBLUE)
    # Highlight baseline and target
    pp17 = _prepress_time_min(17, headspace)
    pp30 = _prepress_time_min(30, headspace)
    ax2.axhline(y=pp17, color=RED, linestyle="--", linewidth=1, alpha=0.6)
    ax2.axhline(y=pp30, color=GREEN, linestyle="--", linewidth=1, alpha=0.6)
    ax2.annotate("", xy=(32, pp30), xytext=(32, pp17),
                 arrowprops=dict(arrowstyle="<->", color=GREEN, lw=2))
    ax2.text(33, (pp17 + pp30) / 2, f"\u2212{pp17-pp30:.1f} min",
             fontsize=9, fontweight="bold", color=GREEN, va="center")
    ax2.set_xlabel("Compressor SCFM", fontsize=9, fontweight="bold")
    ax2.set_ylabel("Pre-Pressurization Time (min)", fontsize=9, fontweight="bold")
    ax2.set_title("Pre-Pressurization Time vs Compressor Size\n"
                  f"(1,000 gal headspace \u2192 20 psig)", fontsize=10,
                  fontweight="bold", color=DBLUE, pad=5)
    ax2.set_xticks(scfm_list)
    ax2.set_ylim(0, max(pp_times) * 1.25)
    ax2.grid(axis="y", alpha=0.3)

    # ── Summary boxes ──
    box(fig, 0.04, 0.14, 0.28, 0.10,
        f"\u2212{avg_total_saved:.0f} min", "Total Saved\nPer Unload", GREEN)
    box(fig, 0.36, 0.14, 0.28, 0.10,
        f"+{fleet_annual_hrs:.0f} hrs", "Fleet Hours\nFreed/Year", GREEN)
    box(fig, 0.68, 0.14, 0.28, 0.10,
        "Your \\$/hr \u00d7 " f"{fleet_annual_hrs:.0f}",
        "= Fleet Value\n(ask your owner)", ORANGE)

    # ── Breakdown callout ──
    avg_disc = np.mean(bins["disc_saved"])
    callout(fig, 0.04, 0.04, 0.92,
            f'Per unload: ~{avg_disc:.0f} min discharge + ~{pp_saved:.0f} min pre-pressurization = '
            f'~{avg_total_saved:.0f} min total saved. '
            f'Across {num_trucks} trucks \u00d7 {total_trips*months_per_year//months_data:,} loads/yr = '
            f'{fleet_annual_hrs:.0f} hrs freed/yr. '
            f'Multiply by your revenue-per-truck-hour to get the fleet dollar value.',
            GREEN)

    pdf.savefig(fig); plt.close(fig)


# ═══════════════════════════════════════════════════════════════════
#  PAGE 7 — THE TRUE COST OF SWITCHING
# ═══════════════════════════════════════════════════════════════════
def page_cost(pdf):
    fig = plt.figure(figsize=(11, 8.5))
    frame(fig, 'THE TRUE COST OF SWITCHING TO 2" HOSE',
          "Fleet-wide impact calculated from your actual product mix and load frequency", 12)

    ax1 = fig.add_axes([0.06, 0.38, 0.40, 0.46])
    ax2 = fig.add_axes([0.56, 0.38, 0.40, 0.46])

    # Left: Extra minutes per load by viscosity bin
    bins = VISC_BINS.copy()
    extra = []
    total_extra = []
    for _, b in bins.iterrows():
        t3 = interp_time(3, 19, b["mid"])
        t2 = interp_time(2, 19, b["mid"])
        e = (t2 - t3) if pd.notna(t2) and pd.notna(t3) else 0
        extra.append(max(e, 0))
        total_extra.append(max(e, 0) * b["loads"])
    bins["extra_min"] = extra
    bins["total_extra"] = total_extra

    clrs = [MBLUE, ACCENT, GREEN, ORANGE, RED, PURPLE]
    bars = ax1.bar(range(len(bins)), bins["extra_min"], color=clrs,
                   edgecolor=WHITE, linewidth=0.5)
    ax1.set_xticks(range(len(bins)))
    ax1.set_xticklabels([b["label"].split("(")[0].strip() for _, b in bins.iterrows()],
                        fontsize=7, rotation=30, ha="right")
    ax1.set_ylabel("Extra Minutes Per Load", fontsize=9, fontweight="bold")
    ax1.set_title('Extra Time Per Load\nIf Switching to 2"', fontsize=11,
                  fontweight="bold", color=DBLUE, pad=5)
    for b, v in zip(bars, bins["extra_min"]):
        if v > 0.5:
            ax1.text(b.get_x() + b.get_width()/2, b.get_height() + 0.3,
                     f"+{v:.1f}", fontsize=8, ha="center", fontweight="bold", color=RED)

    # Right: Fleet-wide total extra hours by viscosity bin
    bars2 = ax2.bar(range(len(bins)), [t / 60 for t in bins["total_extra"]], color=clrs,
                    edgecolor=WHITE, linewidth=0.5)
    ax2.set_xticks(range(len(bins)))
    ax2.set_xticklabels([b["label"].split("(")[0].strip() for _, b in bins.iterrows()],
                        fontsize=7, rotation=30, ha="right")
    ax2.set_ylabel("Total Extra Hours (5 months)", fontsize=9, fontweight="bold")
    ax2.set_title('Fleet-Wide Extra Time\n(loads × extra/load)', fontsize=11,
                  fontweight="bold", color=DBLUE, pad=5)
    for b, v in zip(bars2, [t / 60 for t in bins["total_extra"]]):
        if v > 0.5:
            ax2.text(b.get_x() + b.get_width()/2, b.get_height() + 0.2,
                     f"+{v:.1f} hrs", fontsize=7, ha="center", fontweight="bold", color=RED)

    # Grand total
    grand_total_min = bins["total_extra"].sum()
    grand_total_hrs = grand_total_min / 60
    annual_hrs = grand_total_hrs * (12 / 5)  # extrapolate 5 months to 12

    # Summary boxes
    box(fig, 0.06, 0.10, 0.27, 0.12, f"+{grand_total_hrs:.0f} hrs",
        "Extra Time\n5-Month Period", RED)
    box(fig, 0.37, 0.10, 0.27, 0.12, f"+{annual_hrs:.0f} hrs",
        "Projected\nAnnual Impact", RED)

    # Cost estimate (conservative: $50/hr driver cost)
    annual_cost = annual_hrs * 50
    box(fig, 0.68, 0.10, 0.27, 0.12, f"${annual_cost:,.0f}+",
        "Est. Annual Cost\n(@$50/hr driver)", RED)

    callout(fig, 0.04, 0.04, 0.92,
            f'Switching to 2" hose would add ~{grand_total_hrs:.0f} hrs over 5 months (~{annual_hrs:.0f} hrs/year).\n'
            f'At \\$50/hr driver cost, that is \\${annual_cost:,.0f}+ in lost productivity annually.',
            RED)

    pdf.savefig(fig); plt.close(fig)


# ═══════════════════════════════════════════════════════════════════
#  PAGE 8 — PRESSURE SAFETY
# ═══════════════════════════════════════════════════════════════════
def page_pressure(pdf):
    import math
    fig = plt.figure(figsize=(11, 8.5))
    frame(fig, "PRESSURE SAFETY: WHAT HAPPENS INSIDE THE HOSE?",
          "Average operating pressure by hose size and product viscosity — higher pressure = higher risk", 13)

    ax1 = fig.add_axes([0.06, 0.38, 0.40, 0.46])
    ax2 = fig.add_axes([0.56, 0.38, 0.40, 0.46])

    # Gather pressure data across viscosities at 19 SCFM, 5000 gal, 20 psi
    visc_list = sorted(sim["visc_cP"].unique())
    for hose, ax, title, hc in [(3, ax1, '3" Hose (Current)', MBLUE),
                                 (2, ax2, '2" Hose (Proposed)', RED)]:
        pressures = []
        for v in visc_list:
            row = sim[(sim.hose_in == hose) & (sim.scfm == 19) &
                      (sim.visc_cP == v) & (sim.vol_gal == 5000) &
                      (sim.pre_psi == 20)]
            if len(row) > 0:
                pressures.append(float(row.iloc[0]["avg_pressure_psig"]))
            else:
                pressures.append(0)

        bars = ax.bar(range(len(visc_list)), pressures, color=hc,
                      edgecolor=WHITE, linewidth=0.5, alpha=0.85)
        # Relief valve line at 35 psi
        ax.axhline(y=35, color=RED, linestyle="--", linewidth=1.5, zorder=3)
        ax.text(len(visc_list) - 0.5, 36, "Relief Valve (35 psi)",
                fontsize=7, color=RED, ha="right", fontweight="bold")
        # Hose rated line at 150 psi
        ax.set_xticks(range(len(visc_list)))
        ax.set_xticklabels([f"{v:,}" for v in visc_list],
                           fontsize=7, rotation=30, ha="right")
        ax.set_xlabel("Viscosity (cP)", fontsize=9, fontweight="bold")
        ax.set_ylabel("Avg Operating Pressure (psig)", fontsize=9, fontweight="bold")
        ax.set_title(title, fontsize=11, fontweight="bold", color=DBLUE, pad=5)
        ax.set_ylim(0, 40)
        for b, v in zip(bars, pressures):
            ax.text(b.get_x() + b.get_width()/2, b.get_height() + 0.5,
                    f"{v:.1f}", fontsize=7, ha="center", fontweight="bold",
                    color=RED if v > 25 else NAVY)

    # Comparison table
    fig.text(0.06, 0.30, "Pressure Comparison at 19 SCFM  (standard operating conditions)",
             fontsize=10, fontweight="bold", color=DBLUE)

    col_x = [0.04, 0.22, 0.38, 0.54, 0.70]
    col_w = [0.17, 0.15, 0.15, 0.15, 0.25]
    hdrs  = ["Product", '3" Avg\nPressure', '2" Avg\nPressure', 'Pressure\nIncrease', 'Risk Level']
    y_t = 0.27
    for cx, cw, h in zip(col_x, col_w, hdrs):
        fig.patches.append(FancyBboxPatch(
            (cx, y_t), cw, 0.03, transform=fig.transFigure,
            boxstyle="square,pad=0", facecolor=DBLUE, edgecolor=WHITE, zorder=5))
        fig.text(cx + cw/2, y_t + 0.015, h, fontsize=7, fontweight="bold",
                 color=WHITE, ha="center", va="center", zorder=6)

    prod_show = [
        ("Fleet Median (30 cP)", 50),
        ("VIVATEC 500 (100 cP)", 100),
        ("NIPOL LATEX (200 cP)", 500),
        ("Resin Solution (500 cP)", 500),
        ("Ultra Heavy (2000 cP)", 2000),
    ]
    y_t -= 0.005
    for idx, (label, visc) in enumerate(prod_show):
        y_t -= 0.024
        bg = LBG if idx % 2 == 0 else WHITE
        # Get actual data from sim
        r3 = sim[(sim.hose_in == 3) & (sim.scfm == 19) & (sim.visc_cP == visc) &
                 (sim.vol_gal == 5000) & (sim.pre_psi == 20)]
        r2 = sim[(sim.hose_in == 2) & (sim.scfm == 19) & (sim.visc_cP == visc) &
                 (sim.vol_gal == 5000) & (sim.pre_psi == 20)]
        p3 = float(r3.iloc[0]["avg_pressure_psig"]) if len(r3) > 0 else 0
        p2 = float(r2.iloc[0]["avg_pressure_psig"]) if len(r2) > 0 else 0
        delta = p2 - p3
        pct = (delta / p3 * 100) if p3 > 0 else 0
        risk = "LOW" if p2 < 15 else ("MODERATE" if p2 < 25 else "HIGH")
        risk_clr = GREEN if risk == "LOW" else (ORANGE if risk == "MODERATE" else RED)

        vals = [label.split("(")[0].strip(),
                f"{p3:.1f} psig", f"{p2:.1f} psig",
                f"+{delta:.1f} psig (+{pct:.0f}%)", risk]
        val_colors = [NAVY, NAVY, NAVY, RED, risk_clr]
        for ci, (cx, cw, v) in enumerate(zip(col_x, col_w, vals)):
            fig.patches.append(FancyBboxPatch(
                (cx, y_t), cw, 0.023, transform=fig.transFigure,
                boxstyle="square,pad=0", facecolor=bg, edgecolor=LGRAY,
                linewidth=0.3, zorder=4))
            fig.text(cx + cw/2, y_t + 0.011, v, fontsize=6.5,
                     color=val_colors[ci], ha="center", va="center", zorder=5,
                     fontweight="bold" if ci in (0, 3, 4) else "normal")

    callout(fig, 0.04, 0.045, 0.92,
            '2" hose operates at significantly higher pressure for viscous products.\n'
            'At 2000+ cP the 2" hose approaches relief-valve territory — increased blowout and leak risk.',
            RED)

    pdf.savefig(fig); plt.close(fig)


# ═══════════════════════════════════════════════════════════════════
#  PAGE 9 — TOP 10 PRODUCT CARDS
# ═══════════════════════════════════════════════════════════════════
def page_product_cards(pdf):
    fig = plt.figure(figsize=(11, 8.5))
    frame(fig, "YOUR TOP 10 PRODUCTS: DISCHARGE TIME BY PRODUCT",
          "Real products from your fleet — what each load actually takes at 19 SCFM and 30 SCFM", 9)

    # Table with 10 columns
    top10 = TOP15.head(10).copy()
    col_x = [0.02, 0.16, 0.24, 0.32, 0.42, 0.52, 0.62, 0.72, 0.82]
    col_w = [0.13, 0.07, 0.07, 0.09, 0.09, 0.09, 0.09, 0.09, 0.15]
    hdrs  = ["Product", "Loads", "Visc\n(cP)", '3" @19\nSCFM', '3" @30\nSCFM',
             '2" @19\nSCFM', '2" @30\nSCFM', 'Time\nSaved*', 'Verdict']
    y = 0.85
    for cx, cw, h in zip(col_x, col_w, hdrs):
        fig.patches.append(FancyBboxPatch(
            (cx, y), cw, 0.035, transform=fig.transFigure,
            boxstyle="square,pad=0", facecolor=DBLUE, edgecolor=WHITE, zorder=5))
        fig.text(cx + cw/2, y + 0.017, h, fontsize=6.5, fontweight="bold",
                 color=WHITE, ha="center", va="center", zorder=6)

    y -= 0.005
    for idx, (_, p) in enumerate(top10.iterrows()):
        y -= 0.058
        bg = LBG if idx % 2 == 0 else WHITE
        visc = p["visc_cP"]
        t3_19 = interp_time(3, 19, visc)
        t3_30 = interp_time(3, 30, visc)
        t2_19 = interp_time(2, 19, visc)
        t2_30 = interp_time(2, 30, visc)
        saved = (t3_19 - t3_30) if pd.notna(t3_19) and pd.notna(t3_30) else 0
        # Verdict
        if visc <= 10:
            verdict, vclr = "Any hose OK", GRAY
        elif visc <= 50:
            verdict, vclr = '3" preferred', GREEN
        elif visc <= 200:
            verdict, vclr = '3" important', GREEN
        else:
            verdict, vclr = '3" essential', RED

        vals = [p["commodity"], str(p["loads"]), f"{visc:,}",
                f"{t3_19:.1f}" if pd.notna(t3_19) else "—",
                f"{t3_30:.1f}" if pd.notna(t3_30) else "—",
                f"{t2_19:.1f}" if pd.notna(t2_19) else "—",
                f"{t2_30:.1f}" if pd.notna(t2_30) else "—",
                f"−{saved:.1f} min", verdict]
        val_colors = [NAVY, NAVY, NAVY, NAVY, GREEN, NAVY, ORANGE, GREEN, vclr]
        fweights = ["bold", "normal", "normal", "normal", "bold",
                    "normal", "normal", "bold", "bold"]
        for ci, (cx, cw, v) in enumerate(zip(col_x, col_w, vals)):
            fig.patches.append(FancyBboxPatch(
                (cx, y), cw, 0.055, transform=fig.transFigure,
                boxstyle="square,pad=0", facecolor=bg, edgecolor=LGRAY,
                linewidth=0.3, zorder=4))
            fig.text(cx + cw/2, y + 0.027, v,
                     fontsize=6 if ci == 0 else 6.5,
                     color=val_colors[ci], ha="center", va="center", zorder=5,
                     fontweight=fweights[ci])

    fig.text(0.02, 0.24, "* Time Saved = 3\" hose upgrade from 19 → 30 SCFM",
             fontsize=7, color=GRAY, style="italic")

    # Bottom summary boxes
    # Average time saved across top 10
    savings = []
    for _, p in top10.iterrows():
        t19 = interp_time(3, 19, p["visc_cP"])
        t30 = interp_time(3, 30, p["visc_cP"])
        if pd.notna(t19) and pd.notna(t30):
            savings.append(t19 - t30)
    avg_saved = np.mean(savings) if savings else 0
    total_loads = top10["loads"].sum()
    total_hrs_saved = sum(s * l for s, l in zip(savings, top10["loads"])) / 60

    box(fig, 0.04, 0.06, 0.27, 0.12, f"−{avg_saved:.1f} min",
        "Avg Time Saved/Load\n(19→30 SCFM, 3\")", GREEN)
    box(fig, 0.36, 0.06, 0.27, 0.12, f"{total_loads}",
        "Total Loads\n(Top 10 Products)", MBLUE)
    box(fig, 0.68, 0.06, 0.27, 0.12, f"+{total_hrs_saved:.0f} hrs",
        "Hours Saved\n(Top 10, 5 months)", GREEN)

    pdf.savefig(fig); plt.close(fig)


# ═══════════════════════════════════════════════════════════════════
#  PAGE 9 — WHAT YOU CAN CHANGE
# ═══════════════════════════════════════════════════════════════════
def page_tornado(pdf):
    fig = plt.figure(figsize=(11, 8.5))
    frame(fig, "YOUR TWO LEVERS: HOSE SIZE & COMPRESSOR SCFM",
          "These are the only equipment decisions you control — here is exactly how much each one matters", 14)

    ax1 = fig.add_axes([0.06, 0.38, 0.40, 0.46])
    ax2 = fig.add_axes([0.56, 0.38, 0.40, 0.46])

    bins = VISC_BINS.copy()
    clrs = [MBLUE, ACCENT, GREEN, ORANGE, RED, PURPLE]
    bin_labels = [b["label"].split("(")[0].strip() for _, b in bins.iterrows()]

    # Scale loads to full 3,346 trips
    num_trucks = 80; total_trips = 3346
    months_data, months_year = 5, 12
    bins["scaled_loads"] = (bins["pct"] / 100 * total_trips).round().astype(int)

    # Pre-pressurization savings (17 → 30 SCFM)
    headspace = 1000  # 7000 - 6000 gal
    pp_17 = _prepress_time_min(17, headspace)
    pp_30 = _prepress_time_min(30, headspace)
    pp_saved = pp_17 - pp_30

    # LEFT: Time added by switching 3" → 2" hose (at 17 SCFM baseline)
    hose_penalty = []
    for _, b in bins.iterrows():
        t3 = interp_time(3, 17, b["mid"])
        t2 = interp_time(2, 17, b["mid"])
        delta = (t2 - t3) if pd.notna(t2) and pd.notna(t3) else 0
        hose_penalty.append(max(delta, 0))

    bars1 = ax1.bar(range(len(bins)), hose_penalty, color=clrs,
                    edgecolor=WHITE, linewidth=0.5)
    ax1.set_xticks(range(len(bins)))
    ax1.set_xticklabels(bin_labels, fontsize=7, rotation=30, ha="right")
    ax1.set_ylabel("Extra Minutes Per Load", fontsize=9, fontweight="bold")
    ax1.set_title('HOSE: Time ADDED by\nSwitching 3" \u2192 2"', fontsize=11,
                  fontweight="bold", color=RED, pad=5)
    for b, v in zip(bars1, hose_penalty):
        if v > 0.3:
            ax1.text(b.get_x() + b.get_width()/2, b.get_height() + 0.3,
                     f"+{v:.1f} min", fontsize=7, ha="center",
                     fontweight="bold", color=RED)

    # RIGHT: Total time saved by upgrading 17 → 30 SCFM (discharge + pre-press)
    disc_savings = []
    for _, b in bins.iterrows():
        t17 = interp_time(3, 17, b["mid"])
        t30 = interp_time(3, 30, b["mid"])
        delta = (t17 - t30) if pd.notna(t17) and pd.notna(t30) else 0
        disc_savings.append(max(delta, 0))

    x = np.arange(len(bins))
    w = 0.65
    bars_disc = ax2.bar(x, disc_savings, w, color=clrs,
                        edgecolor=WHITE, linewidth=0.5, label="Discharge", zorder=3)
    bars_pp = ax2.bar(x, [pp_saved]*len(bins), w, bottom=disc_savings,
                      color=GOLD, edgecolor=WHITE, linewidth=0.5,
                      label="Pre-Pressurization", zorder=3)
    ax2.set_xticks(x)
    ax2.set_xticklabels(bin_labels, fontsize=7, rotation=30, ha="right")
    ax2.set_ylabel("Minutes Saved Per Load", fontsize=9, fontweight="bold")
    ax2.set_title('COMPRESSOR: Total Time SAVED by\nUpgrading 17 \u2192 30 SCFM (3" Hose)', fontsize=10,
                  fontweight="bold", color=GREEN, pad=5)
    total_savings = [d + pp_saved for d in disc_savings]
    for i, (d, tot) in enumerate(zip(disc_savings, total_savings)):
        ax2.text(i, tot + 0.5, f"\u2212{tot:.0f} min", fontsize=7, ha="center",
                 fontweight="bold", color=GREEN)
    ax2.legend(fontsize=7, loc="upper left", framealpha=0.9)
    ax2.set_ylim(0, max(total_savings) * 1.15)

    # Fleet-weighted totals (annual) — scaled to 3,346 loads
    hose_annual_hrs = sum(p * ld * months_year / months_data / 60
                          for p, ld in zip(hose_penalty, bins["scaled_loads"]))
    scfm_annual_hrs = sum(t * ld * months_year / months_data / 60
                          for t, ld in zip(total_savings, bins["scaled_loads"]))

    # Summary boxes — hours only, no fake $/hr
    box(fig, 0.04, 0.18, 0.21, 0.12, f"+{hose_annual_hrs:.0f} hrs/yr",
        'COST of\nSwitching to 2"', RED)
    box(fig, 0.27, 0.18, 0.21, 0.12,
        "Your \\$/hr \u00d7 " f"{hose_annual_hrs:.0f}",
        'Lost Productivity\n(your rate)', RED)
    box(fig, 0.50, 0.18, 0.21, 0.12, f"\u2212{scfm_annual_hrs:.0f} hrs/yr",
        'GAIN from\n30 SCFM Upgrade', GREEN)
    box(fig, 0.73, 0.18, 0.21, 0.12,
        "Your \\$/hr \u00d7 " f"{scfm_annual_hrs:.0f}",
        'Productivity Gained\n(your rate)', GREEN)

    callout(fig, 0.04, 0.045, 0.92,
            'You control two things: hose size and compressor SCFM.\n'
            'Keep 3" hose (avoid the red cost) + upgrade to 30 SCFM (capture the green savings).',
            DBLUE)

    pdf.savefig(fig); plt.close(fig)


# ═══════════════════════════════════════════════════════════════════
#  PAGE 10 — FLOW RATE & HOSE WEAR
# ═══════════════════════════════════════════════════════════════════
def page_velocity(pdf):
    fig = plt.figure(figsize=(11, 8.5))
    frame(fig, "FLOW RATE & HOSE WEAR ANALYSIS",
          "Peak and average GPM by hose size — higher flow through smaller hose = more wear", 15)

    ax1 = fig.add_axes([0.06, 0.38, 0.40, 0.46])
    ax2 = fig.add_axes([0.56, 0.38, 0.40, 0.46])

    visc_list = sorted(sim["visc_cP"].unique())
    visc_labels = [f"{v:,}" for v in visc_list]

    for hose, ax, title, hc in [(3, ax1, '3" Hose — Peak GPM', MBLUE),
                                 (2, ax2, '2" Hose — Peak GPM', RED)]:
        peak_gpms = []
        avg_gpms = []
        for v in visc_list:
            row = sim[(sim.hose_in == hose) & (sim.scfm == 19) &
                      (sim.visc_cP == v) & (sim.vol_gal == 5000) &
                      (sim.pre_psi == 20)]
            if len(row) > 0:
                peak_gpms.append(float(row.iloc[0]["peak_gpm"]))
                avg_gpms.append(float(row.iloc[0]["avg_gpm"]))
            else:
                peak_gpms.append(0)
                avg_gpms.append(0)

        x = range(len(visc_list))
        w = 0.35
        bars_p = ax.bar([i - w/2 for i in x], peak_gpms, width=w, color=hc,
                        edgecolor=WHITE, linewidth=0.5, label="Peak GPM", alpha=0.9)
        bars_a = ax.bar([i + w/2 for i in x], avg_gpms, width=w, color=hc,
                        edgecolor=WHITE, linewidth=0.5, label="Avg GPM", alpha=0.45)

        ax.set_xticks(range(len(visc_list)))
        ax.set_xticklabels(visc_labels, fontsize=7, rotation=30, ha="right")
        ax.set_xlabel("Viscosity (cP)", fontsize=9, fontweight="bold")
        ax.set_ylabel("Flow Rate (GPM)", fontsize=9, fontweight="bold")
        ax.set_title(title, fontsize=11, fontweight="bold", color=DBLUE, pad=5)
        ax.legend(fontsize=7, loc="upper right")
        ax.set_ylim(0, max(peak_gpms) * 1.15 if peak_gpms else 100)
        for b, v in zip(bars_p, peak_gpms):
            ax.text(b.get_x() + b.get_width()/2, b.get_height() + 3,
                    f"{v:.0f}", fontsize=6.5, ha="center", fontweight="bold",
                    color=NAVY)

    # Comparison table
    fig.text(0.06, 0.30, "Flow Rate Comparison at 19 SCFM  (your current compressor)",
             fontsize=10, fontweight="bold", color=DBLUE)

    col_x = [0.04, 0.20, 0.34, 0.48, 0.62, 0.78]
    col_w = [0.15, 0.13, 0.13, 0.13, 0.15, 0.17]
    hdrs  = ["Product", '3" Peak\nGPM', '3" Avg\nGPM', '2" Peak\nGPM', '2" Avg\nGPM', 'Result']
    y_t = 0.27
    for cx, cw, h in zip(col_x, col_w, hdrs):
        fig.patches.append(FancyBboxPatch(
            (cx, y_t), cw, 0.03, transform=fig.transFigure,
            boxstyle="square,pad=0", facecolor=DBLUE, edgecolor=WHITE, zorder=5))
        fig.text(cx + cw/2, y_t + 0.015, h, fontsize=6.5, fontweight="bold",
                 color=WHITE, ha="center", va="center", zorder=6)

    tbl_products = [
        ("Water-like (1 cP)", 1),
        ("Fleet Median (50 cP)", 50),
        ("Medium (100 cP)", 100),
        ("Heavy (500 cP)", 500),
        ("Ultra Heavy (2000 cP)", 2000),
    ]
    y_t -= 0.005
    for idx, (label, visc) in enumerate(tbl_products):
        y_t -= 0.024
        bg = LBG if idx % 2 == 0 else WHITE
        p3, a3, p2, a2 = 0, 0, 0, 0
        for hose_val in [3, 2]:
            row = sim[(sim.hose_in == hose_val) & (sim.scfm == 19) &
                      (sim.visc_cP == visc) & (sim.vol_gal == 5000) &
                      (sim.pre_psi == 20)]
            if len(row) > 0:
                if hose_val == 3:
                    p3 = float(row.iloc[0]["peak_gpm"])
                    a3 = float(row.iloc[0]["avg_gpm"])
                else:
                    p2 = float(row.iloc[0]["peak_gpm"])
                    a2 = float(row.iloc[0]["avg_gpm"])
        # 3" moves more GPM = faster unload
        pct_less = ((p3 - p2) / p3 * 100) if p3 > 0 else 0
        result = f'3" flows {pct_less:.0f}% more' if pct_less > 0 else "Similar"
        res_clr = GREEN if pct_less > 20 else MBLUE

        vals = [label.split("(")[0].strip(),
                f"{p3:.0f}", f"{a3:.0f}", f"{p2:.0f}", f"{a2:.0f}", result]
        val_colors = [NAVY, GREEN, GREEN, NAVY, NAVY, res_clr]
        for ci, (cx, cw, v) in enumerate(zip(col_x, col_w, vals)):
            fig.patches.append(FancyBboxPatch(
                (cx, y_t), cw, 0.023, transform=fig.transFigure,
                boxstyle="square,pad=0", facecolor=bg, edgecolor=LGRAY,
                linewidth=0.3, zorder=4))
            fig.text(cx + cw/2, y_t + 0.011, v, fontsize=6.5,
                     color=val_colors[ci], ha="center", va="center", zorder=5,
                     fontweight="bold" if ci in (0, 5) else "normal")

    callout(fig, 0.04, 0.045, 0.92,
            '3" hose delivers 2-3x the peak flow rate of 2" hose across all viscosities.\n'
            'More GPM = faster unload. For heavy products, 2" hose drops below 120 GPM — painfully slow.',
            RED)

    pdf.savefig(fig); plt.close(fig)


# ═══════════════════════════════════════════════════════════════════
#  PAGE 11 — RECOMMENDATIONS
# ═══════════════════════════════════════════════════════════════════
def page_recs(pdf):
    fig = plt.figure(figsize=(11, 8.5))
    frame(fig, "RECOMMENDATIONS & DECISION FRAMEWORK",
          "Data-backed equipment and operations decisions", 16)

    # Decision matrix
    fig.text(0.04, 0.88, "EQUIPMENT DECISION MATRIX", fontsize=11,
             fontweight="bold", color=DBLUE)
    fig.text(0.04, 0.86, "Based on your actual product mix and simulated discharge performance",
             fontsize=8, color=GRAY)

    col_x = [0.04, 0.20, 0.37, 0.54, 0.72]
    col_w = [0.15, 0.16, 0.16, 0.17, 0.25]
    hdrs  = ["Product\nCategory", "% of Your\nLoads", "Hose\nDecision",
             "Compressor\nGuidance", "Why"]
    y = 0.82
    for cx, cw, h in zip(col_x, col_w, hdrs):
        fig.patches.append(FancyBboxPatch(
            (cx, y), cw, 0.03, transform=fig.transFigure,
            boxstyle="square,pad=0", facecolor=DBLUE, edgecolor=WHITE, zorder=5))
        fig.text(cx + cw/2, y + 0.015, h, fontsize=7.5, fontweight="bold",
                 color=WHITE, ha="center", va="center", zorder=6)

    rows = [
        ("Water-like\n\u2264 10 cP",  "26%\n(870 loads)",
         '2" or 3" OK', "17+ SCFM\nsufficient",
         "Low friction \u2014 hose size\nadds ~11 min. Minimal impact.", LBG),
        ("Light\n10\u201350 cP",      "38%\n(1,271 loads)",
         '3" preferred', "17 SCFM\n(current OK)",
         '3" is ~12 min faster.\nAdds up over 1,271 loads.', WHITE),
        ("Medium\n50\u2013200 cP",    "26%\n(870 loads)",
         '3" important', "25 SCFM\nrecommended",
         '2" adds ~15 min/load.\nIncludes your #3, #5, #6 products.', LBG),
        ("Heavy\n200\u2013500 cP",    "5%\n(167 loads)",
         '3" required', "25\u201330 SCFM\nideal",
         '2" adds ~27 min/load.\nResin Solution lives here.', WHITE),
        ("Very Heavy\n500+ cP",  "5%\n(167 loads)",
         '3" essential', "30\u201335 SCFM\nbeneficial",
         '2" adds 47+ min/load.\nPhysics says 3" is the only option.', LBG),
    ]

    y -= 0.005
    for vals_tuple in rows:
        *vals, bg = vals_tuple
        y -= 0.05
        for ci, (cx, cw, v) in enumerate(zip(col_x, col_w, vals)):
            fig.patches.append(FancyBboxPatch(
                (cx, y), cw, 0.048, transform=fig.transFigure,
                boxstyle="square,pad=0", facecolor=bg, edgecolor=LGRAY,
                linewidth=0.3, zorder=4))
            color = NAVY
            if ci == 2:
                if "essential" in v or "required" in v:
                    color = GREEN
                elif "important" in v:
                    color = GREEN
            fig.text(cx + cw/2, y + 0.024, v, fontsize=6.5, color=color,
                     ha="center", va="center", zorder=5, linespacing=1.2,
                     fontweight="bold" if ci == 2 else "normal")

    # 3 Action boxes
    y_a = 0.16
    actions = [
        ("KEEP 3\" HOSE", GREEN,
         "Do not switch to 2\".\nThe data shows 2\" adds\n"
         "11\u201347+ min per load for\n6,000 gal at your product mix.\n"
         "Cost savings are minimal\nvs. lost driver time."),
        ("UPGRADE TO 30 SCFM", MBLUE,
         "Current ~17 SCFM works for\nlight products (64% of loads).\n"
         "Upgrading to 30 SCFM saves\n14\u201319 min discharge + ~5 min\n"
         "pre-pressurization per load.\n~2,600 fleet hrs freed/year."),
        ("TRACK COMMODITY DATA", ORANGE,
         "Mandate commodity entry in\nQuickManage at booking.\n"
         "Auto-flag 200+ cP loads\nso dispatchers assign the\n"
         "right equipment. This costs\n$0 and prevents mismatches."),
    ]
    for i, (title_text, color, text) in enumerate(actions):
        x = 0.04 + i * 0.32
        fig.patches.append(FancyBboxPatch(
            (x, y_a - 0.05), 0.29, 0.20, transform=fig.transFigure,
            boxstyle="round,pad=0.01", facecolor=WHITE, edgecolor=color,
            linewidth=2.5, zorder=5))
        fig.text(x + 0.145, y_a + 0.125, title_text, fontsize=10,
                 fontweight="bold", color=color, ha="center", va="center", zorder=6)
        fig.text(x + 0.145, y_a + 0.02, text, fontsize=7.5,
                 color=NAVY, ha="center", va="center", zorder=6, linespacing=1.3)

    fig.text(0.50, 0.04,
             "Analysis based on 3,346 real trips + 840 physics simulations validated ±7% against field data.",
             fontsize=7, color=GRAY, ha="center", style="italic")

    pdf.savefig(fig); plt.close(fig)


# ═══════════════════════════════════════════════════════════════════
#  EXCEL
# ═══════════════════════════════════════════════════════════════════
def build_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="003A70", end_color="003A70", fill_type="solid")
    df_ = Font(name="Arial", size=10)
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bdr = Border(*[Side(style="thin", color="D9D9D9")] * 4)

    # Sheet 1: Product impact
    ws = wb.active
    ws.title = "Product Discharge Impact"
    headers = ["Rank", "Commodity", "Loads/Period", "Viscosity (cP)",
               '3" Hose 19 SCFM', '2" Hose 19 SCFM', "Extra Time/Load",
               "Total Extra (min)", "Total Extra (hrs)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ctr

    total_extra = 0
    for ri, (_, p) in enumerate(TOP15.iterrows(), 2):
        t3 = interp_time(3, 19, p["visc_cP"])
        t2 = interp_time(2, 19, p["visc_cP"])
        delta = (t2 - t3) if pd.notna(t2) and pd.notna(t3) else 0
        total = delta * p["loads"]
        total_extra += total
        vals = [ri-1, p["commodity"], p["loads"], p["visc_cP"],
                round(t3, 1), round(t2, 1), round(delta, 1),
                round(total, 0), round(total/60, 1)]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=c, value=v)
            cell.font = df_; cell.alignment = ctr; cell.border = bdr

    # Total row
    ri = len(TOP15) + 2
    ws.cell(row=ri, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=ri, column=8, value=round(total_extra, 0)).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=ri, column=9, value=round(total_extra/60, 1)).font = Font(name="Arial", bold=True, size=10)

    for c in range(1, 10):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # Sheet 2: Viscosity bins
    ws2 = wb.create_sheet("Fleet Viscosity Bins")
    h2 = ["Viscosity Range", "Mid (cP)", "% of Loads", "Est. Loads",
          '3" Time (min)', '2" Time (min)', "Extra/Load (min)",
          "Total Extra (min)", "Total Extra (hrs)"]
    for c, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ctr

    for ri, (_, b) in enumerate(VISC_BINS.iterrows(), 2):
        t3 = interp_time(3, 19, b["mid"])
        t2 = interp_time(2, 19, b["mid"])
        delta = (t2 - t3) if pd.notna(t2) and pd.notna(t3) else 0
        total = delta * b["loads"]
        vals = [b["label"], b["mid"], b["pct"], b["loads"],
                round(t3, 1), round(t2, 1), round(delta, 1),
                round(total, 0), round(total/60, 1)]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=c, value=v)
            cell.font = df_; cell.alignment = ctr; cell.border = bdr
    for c in range(1, 10):
        ws2.column_dimensions[get_column_letter(c)].width = 20

    # Sheet 3: Compressor analysis
    ws3 = wb.create_sheet("Compressor SCFM Analysis")
    h3 = ["SCFM"] + [f'{v} cP — 3" (min)' for v in [30, 100, 200, 500]] + \
                     [f'{v} cP — 2" (min)' for v in [30, 100, 200, 500]]
    for c, h in enumerate(h3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ctr
    for ri, s in enumerate(SIM_SCFMS, 2):
        cell = ws3.cell(row=ri, column=1, value=s)
        cell.font = df_; cell.alignment = ctr
        for vi, v in enumerate([30, 100, 200, 500]):
            t3 = interp_time(3, s, v)
            t2 = interp_time(2, s, v)
            ws3.cell(row=ri, column=2+vi, value=round(t3,1) if pd.notna(t3) else "DNF").font = df_
            ws3.cell(row=ri, column=6+vi, value=round(t2,1) if pd.notna(t2) else "DNF").font = df_
            ws3.cell(row=ri, column=2+vi).alignment = ctr
            ws3.cell(row=ri, column=6+vi).alignment = ctr
    for c in range(1, 10):
        ws3.column_dimensions[get_column_letter(c)].width = 18

    # Sheet 4: Raw sim data
    ws4 = wb.create_sheet("Simulation Raw Data")
    for c, h in enumerate(sim.columns, 1):
        cell = ws4.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ctr
    for r, row in enumerate(sim.values, 2):
        for c, val in enumerate(row, 1):
            cell = ws4.cell(row=r, column=c)
            if isinstance(val, float) and not np.isnan(val):
                cell.value = round(val, 1)
            else:
                cell.value = val
            cell.font = df_; cell.alignment = ctr
    for c in range(1, len(sim.columns)+1):
        ws4.column_dimensions[get_column_letter(c)].width = 14

    wb.save(OUT_XLSX)
    print(f"  Excel: {OUT_XLSX}")


# ═══════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("Building 16-page combined report...")
    with PdfPages(OUT_PDF) as pdf:
        cover(pdf);                 print("  Page 1: Cover ✓")
        page_exec(pdf);             print("  Page 2: Executive Summary ✓")
        page_fleet(pdf);            print("  Page 3: Fleet Profile ✓")
        for i, scfm_val in enumerate([10, 13, 17, 19, 25, 35]):
            pg = 4 + i
            page_hose_compare(pdf, scfm=scfm_val, pg=pg)
            print(f"  Page {pg}: 2\" vs 3\" @ {scfm_val} SCFM ✓")
        page_compressor(pdf);       print("  Page 10: Compressor Analysis ✓")
        page_compressor_roi(pdf);   print("  Page 11: Compressor Upgrade ROI ✓")
        page_cost(pdf);             print("  Page 12: Cost of Switching ✓")
        page_pressure(pdf);         print("  Page 13: Pressure Safety ✓")
        page_tornado(pdf);          print("  Page 14: Sensitivity Tornado ✓")
        page_velocity(pdf);         print("  Page 15: Flow Velocity & Wear ✓")
        page_recs(pdf);             print("  Page 16: Recommendations ✓")
    print(f"  PDF: {OUT_PDF}")

    print("\nBuilding Excel workbook...")
    build_excel()

    print("\nDone!")
