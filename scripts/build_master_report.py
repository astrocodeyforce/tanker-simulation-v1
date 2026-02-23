#!/usr/bin/env python3
"""Build Master Excel Report — All Sweeps A-H in one workbook."""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = Workbook()

# ── Style definitions ──
hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="2F5496")
num_font = Font(size=10)
title_font = Font(bold=True, size=14, color="1F3864")
subtitle_font = Font(bold=True, size=11, color="2F5496")
good_fill = PatternFill("solid", fgColor="C6EFCE")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))

def style_header(ws, row=1, ncols=20):
    for col in range(1, ncols+1):
        cell = ws.cell(row=row, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def style_data(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row+1):
        for c in range(1, ncols+1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = num_font
            cell.alignment = Alignment(horizontal='center')

def auto_width(ws, ncols, min_w=10, max_w=22):
    for c in range(1, ncols+1):
        ws.column_dimensions[get_column_letter(c)].width = min(max_w, max(min_w,
            max((len(str(ws.cell(row=r, column=c).value or ''))
                 for r in range(1, ws.max_row+1)), default=10) + 2))

def add_data_sheet(wb, name, df, title):
    ws = wb.create_sheet(title=name)
    ws.cell(row=1, column=1, value=title).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=min(len(df.columns), 8))
    ws.cell(row=2, column=1, value=f"{len(df)} simulations").font = subtitle_font
    for j, col in enumerate(df.columns, 1):
        ws.cell(row=4, column=j, value=col)
    style_header(ws, row=4, ncols=len(df.columns))
    for i, (_, row) in enumerate(df.iterrows(), 5):
        for j, col in enumerate(df.columns, 1):
            val = row[col]
            cell = ws.cell(row=i, column=j, value=val)
            if isinstance(val, float):
                cell.number_format = '0.00'
    style_data(ws, 5, 4+len(df), len(df.columns))
    auto_width(ws, len(df.columns))
    ws.freeze_panes = 'A5'
    return ws

# ═══════════════════════════════════════════════
# Load all data
# ═══════════════════════════════════════════════
BASE = '/work/data/parametric_sweeps'
A = pd.read_csv(f'{BASE}/A_liquid_volume.csv')
B = pd.read_csv(f'{BASE}/B_tank_pressure.csv')
C = pd.read_csv(f'{BASE}/C_gas_temperature.csv')
D = pd.read_csv(f'{BASE}/D_compressor_scfm.csv')
E = pd.read_csv(f'{BASE}/E_hose_diameter_multi.csv')
F = pd.read_csv(f'{BASE}/F_viscosity.csv')
G = pd.read_csv(f'{BASE}/G_visc_scfm_combo.csv')
H = pd.read_csv(f'{BASE}/H_mega_combo.csv')

total_sims = sum(len(x) for x in [A,B,C,D,E,F,G,H])
print(f"Loaded {total_sims} total simulation rows")

# ═══════════════════════════════════════════════
# Sheet 0: EXECUTIVE SUMMARY
# ═══════════════════════════════════════════════
ws0 = wb.active
ws0.title = "Executive Summary"

r = 1
ws0.cell(row=r, column=1, value="TANKER TRANSFER SIMULATION \u2014 MASTER REPORT").font = title_font
ws0.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
r = 2
ws0.cell(row=r, column=1, value=f"{total_sims:,} OpenModelica Simulations | Sweeps A\u2013H").font = subtitle_font
r = 4
ws0.cell(row=r, column=1, value="VARIABLE IMPORTANCE (from 1080-sim ANOVA)").font = subtitle_font
r = 6

# Variable importance table
vi_headers = ["Rank", "Factor", "Variance %", "Spread (min)", "Controllable?", "Action"]
for j, h in enumerate(vi_headers, 1):
    ws0.cell(row=r, column=j, value=h)
style_header(ws0, row=r, ncols=6)

vi_data = [
    ("#1", "Viscosity", "32.4%", "28.4", "No", "Accept \u2014 product-dependent"),
    ("#2", "Compressor SCFM", "30.1%", "26.8", "YES", "UPGRADE FIRST"),
    ("#3", "Hose Diameter", "19.8%", "18.8", "YES", "Pair with compressor"),
    ("#4", "Liquid Volume", "3.6%", "8.7", "No", "Load size varies"),
    ("#5", "Initial Pressure", "0.2%", "2.0", "Yes", "DON\u2019T BOTHER \u2014 0 psig always wins"),
]
for row_data in vi_data:
    r += 1
    for j, val in enumerate(row_data, 1):
        cell = ws0.cell(row=r, column=j, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

r += 2
ws0.cell(row=r, column=1, value="PRESSURE VERDICT").font = subtitle_font
r += 1
ws0.cell(row=r, column=1, value="Pre-pressurizing HURTS in all 1080 combinations tested. 0 psig is ALWAYS fastest. Open valve immediately.")
r += 2
ws0.cell(row=r, column=1, value="TOP 3 SOLUTIONS (6000 gal, 0 psig, SG 1.34)").font = subtitle_font

# Solution comparison table
r += 2
ws0.cell(row=r, column=1, value="Solution Comparison at 6000 gal (minutes)").font = subtitle_font
r += 1
sol_headers = ["Config", "Equipment", "Cost", "1 cP", "10 cP", "100 cP", "500 cP", "1000 cP", "2000 cP"]
for j, h in enumerate(sol_headers, 1):
    ws0.cell(row=r, column=j, value=h)
style_header(ws0, row=r, ncols=len(sol_headers))

sol_configs = [
    ("CURRENT",    '3" / 19 SCFM', "$0",      3, 19, 0),
    ("SOLUTION 1", '3" / 35 SCFM', "$2-4K",   3, 35, 0),
    ("SOLUTION 2", '4" / 35 SCFM', "$3.5-6K", 4, 35, 0),
    ("SOLUTION 3", '4" / 50 SCFM', "$5-9K",   4, 50, 0),
]

for sol_label, equip, cost, diam, scfm, press in sol_configs:
    r += 1
    ws0.cell(row=r, column=1, value=sol_label).border = thin_border
    ws0.cell(row=r, column=2, value=equip).border = thin_border
    ws0.cell(row=r, column=3, value=cost).border = thin_border
    for j, visc in enumerate([1, 10, 100, 500, 1000, 2000], 4):
        sub = H[(H['viscosity_cP']==visc) & (H['compressor_scfm']==scfm) &
                (H['hose_diameter_in']==diam) & (H['liquid_volume_gal']==6000) &
                (H['initial_pressure_psig']==press)]
        if len(sub) > 0:
            val = round(sub.iloc[0]['total_time_min'], 1)
            cell = ws0.cell(row=r, column=j, value=val)
            cell.number_format = '0.0'
        else:
            cell = ws0.cell(row=r, column=j, value="N/A")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

# % improvement rows
r += 2
ws0.cell(row=r, column=1, value="% Improvement vs Current").font = subtitle_font
r += 1
for sol_label, equip, cost, diam, scfm, press in sol_configs[1:]:
    ws0.cell(row=r, column=1, value=sol_label).border = thin_border
    ws0.cell(row=r, column=2, value=equip).border = thin_border
    for j, visc in enumerate([1, 10, 100, 500, 1000, 2000], 4):
        base = H[(H['viscosity_cP']==visc) & (H['compressor_scfm']==19) &
                 (H['hose_diameter_in']==3) & (H['liquid_volume_gal']==6000) &
                 (H['initial_pressure_psig']==0)]
        sol = H[(H['viscosity_cP']==visc) & (H['compressor_scfm']==scfm) &
                (H['hose_diameter_in']==diam) & (H['liquid_volume_gal']==6000) &
                (H['initial_pressure_psig']==press)]
        if len(base)>0 and len(sol)>0:
            pct = (base.iloc[0]['total_time_min'] - sol.iloc[0]['total_time_min']) / base.iloc[0]['total_time_min'] * 100
            cell = ws0.cell(row=r, column=j, value=f"+{pct:.1f}%")
            cell.fill = good_fill
        else:
            cell = ws0.cell(row=r, column=j, value="N/A")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    r += 1

# All volumes table
r += 1
ws0.cell(row=r, column=1, value="Comparison Across All Volumes (0 psig, minutes)").font = subtitle_font
r += 1
vol_hdrs = ["Volume", "Config", "1 cP", "100 cP", "500 cP", "1000 cP", "2000 cP"]
for j, h in enumerate(vol_hdrs, 1):
    ws0.cell(row=r, column=j, value=h)
style_header(ws0, row=r, ncols=len(vol_hdrs))

for vol in [4000, 5000, 6000, 6500]:
    for label, diam, scfm in [("Current 3\"/19", 3, 19), ("Sol 1 3\"/35", 3, 35),
                               ("Sol 2 4\"/35", 4, 35), ("Sol 3 4\"/50", 4, 50)]:
        r += 1
        ws0.cell(row=r, column=1, value=f"{vol} gal").border = thin_border
        ws0.cell(row=r, column=2, value=label).border = thin_border
        for j, visc in enumerate([1, 100, 500, 1000, 2000], 3):
            sub = H[(H['viscosity_cP']==visc) & (H['compressor_scfm']==scfm) &
                    (H['hose_diameter_in']==diam) & (H['liquid_volume_gal']==vol) &
                    (H['initial_pressure_psig']==0)]
            if len(sub) > 0:
                cell = ws0.cell(row=r, column=j, value=round(sub.iloc[0]['total_time_min'], 1))
                cell.number_format = '0.0'
            else:
                cell = ws0.cell(row=r, column=j, value="N/A")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

# Pressure detail table
r += 2
ws0.cell(row=r, column=1, value="Pressure Effect Detail (6000 gal, 3\" hose)").font = subtitle_font
r += 1
pr_hdrs = ["SCFM", "Viscosity", "0 psig (total)", "10 psig (total)", "22 psig (total)", "Best"]
for j, h in enumerate(pr_hdrs, 1):
    ws0.cell(row=r, column=j, value=h)
style_header(ws0, row=r, ncols=len(pr_hdrs))

for scfm in [19, 35, 50, 64]:
    for visc in [1, 100, 500, 1000, 2000]:
        r += 1
        ws0.cell(row=r, column=1, value=scfm).border = thin_border
        ws0.cell(row=r, column=2, value=f"{visc} cP").border = thin_border
        times = {}
        for j, press in enumerate([0, 10, 22], 3):
            sub = H[(H['viscosity_cP']==visc) & (H['compressor_scfm']==scfm) &
                    (H['hose_diameter_in']==3) & (H['liquid_volume_gal']==6000) &
                    (H['initial_pressure_psig']==press)]
            if len(sub) > 0:
                t = round(sub.iloc[0]['total_time_min'], 1)
                pt = round(sub.iloc[0]['pressurize_time_min'], 1)
                vt = round(sub.iloc[0]['valve_transfer_time_min'], 1)
                cell = ws0.cell(row=r, column=j, value=f"{t} ({pt}+{vt})")
                times[press] = t
            else:
                cell = ws0.cell(row=r, column=j, value="N/A")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        best_press = min(times, key=times.get) if times else 0
        cell = ws0.cell(row=r, column=6, value=f"{best_press} psig")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if best_press == 0:
            cell.fill = good_fill

# Sweep inventory
r += 2
ws0.cell(row=r, column=1, value="SWEEP INVENTORY").font = subtitle_font
r += 1
inv_hdrs = ["Sweep", "Variable(s)", "Sims", "Key Finding"]
for j, h in enumerate(inv_hdrs, 1):
    ws0.cell(row=r, column=j, value=h)
style_header(ws0, row=r, ncols=4)

inventory = [
    ("A", "Liquid Volume (4000-6500 gal)", len(A), "+1000 gal = +6.9 min (linear)"),
    ("B", "Tank Pressure (0-24.5 psig)", len(B), "Only 11% impact \u2014 break-even ROI"),
    ("C", "Gas Temperature (15-40\u00b0C)", len(C), "Negligible \u2014 6.7% total impact"),
    ("D", "Compressor SCFM (15-64)", len(D), "DOMINANT \u2014 62.6% reduction, sweet spot 35"),
    ("E", "Hose Diam \u00d7 Vol \u00d7 SCFM", len(E), "Diameter + compressor are LINKED"),
    ("F", "Viscosity (1-2000 cP)", len(F), "#1 variable \u2014 134% impact range"),
    ("G", "Viscosity \u00d7 Compressor", len(G), "2000 cP wall \u2014 compressor ROI collapses"),
    ("H", "5D Mega (visc\u00d7scfm\u00d7diam\u00d7vol\u00d7press)", len(H), "0 psig always wins. Pressure is useless."),
]
for sweep, var, sims, finding in inventory:
    r += 1
    ws0.cell(row=r, column=1, value=sweep).border = thin_border
    ws0.cell(row=r, column=2, value=var).border = thin_border
    ws0.cell(row=r, column=3, value=sims).border = thin_border
    ws0.cell(row=r, column=4, value=finding).border = thin_border

r += 1
ws0.cell(row=r, column=1, value=f"TOTAL: {total_sims:,} simulations").font = subtitle_font

auto_width(ws0, 9, min_w=12, max_w=45)

# ═══════════════════════════════════════════════
# Data sheets A-H
# ═══════════════════════════════════════════════
add_data_sheet(wb, "A - Liquid Volume", A, "Sweep A \u2014 Liquid Volume (4000-6500 gal)")
add_data_sheet(wb, "B - Tank Pressure", B, "Sweep B \u2014 Tank Pressure (0-24.5 psig)")
add_data_sheet(wb, "C - Gas Temperature", C, "Sweep C \u2014 Gas Temperature (15-40\u00b0C)")
add_data_sheet(wb, "D - Compressor SCFM", D, "Sweep D \u2014 Compressor SCFM (15-64)")
add_data_sheet(wb, "E - Hose Diameter", E, "Sweep E \u2014 Hose Diameter \u00d7 Volume \u00d7 SCFM (100 combos)")
add_data_sheet(wb, "F - Viscosity", F, "Sweep F \u2014 Viscosity (1-2000 cP)")
add_data_sheet(wb, "G - Visc x Compressor", G, "Sweep G \u2014 Viscosity \u00d7 Compressor (10\u00d710 grid)")
add_data_sheet(wb, "H - Mega 5D Combo", H, "Sweep H \u2014 5D Mega Combo (1080 sims)")

# ═══════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════
outpath = '/work/data/parametric_sweeps/Master_Sweep_Report.xlsx'
wb.save(outpath)
size = os.path.getsize(outpath)
print(f"Saved: {outpath} ({size:,} bytes)")
print(f"Sheets: {wb.sheetnames}")
print(f"Total: {total_sims:,} simulations across {len(wb.sheetnames)} sheets")
